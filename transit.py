import pandas as pd
import keyboard
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from users import User
import getpass

namesList = []
dniList = []
teorico = []
practico = []
guardado = ""

def UsrEnter():
    try:
        leave = " "
        print("\nBienvenido al sistema de Almirante Brown.\n")
        print("\n  1. Iniciar sesión.\n  2. Crear cuenta.\n  3. Salir del sistema.")
        while True:
            if keyboard.is_pressed('1'):
                while True:
                    his_id = input("\nIngrese su nombre de usuario: ")
                    his_pass = getpass.getpass(prompt="Ingrese su contraseña: ")
                    if his_id != "" and User.user_exists(his_id, his_pass):
                        leave = MenuPrincipal()
                        break
                    else:
                        print("Nombre de usuario o contraseña incorrectos.")
                        continue
            elif keyboard.is_pressed('2'):
                createUsr()
                break
            elif keyboard.is_pressed('3'):
                raise KeyboardInterrupt      
            if leave == "y":
                break
            else:
                continue
    except KeyboardInterrupt:
        print("\nUsted ha salido del sistema de Almirante Brown. Vuelva pronto")


def createUsr():
    try:
        while True:
            usrId = input("\n\n\nNombre de usuario: ")
            usrPass = input("Contraseña: ")
            usrName = input("Nombre: ")
            usrSurname = input("Apellido: ")
            usrRole = input("Eres inspector? ").lower()
            if usrRole == "si" or usrRole == "y":
                usrRole = "admin"
            else:
                usrRole = " "

            newuser = User(usrId, usrPass, usrName, usrSurname, usrRole)
            check_if_saved = newuser.save_usr()
            if check_if_saved:
                print("Usuario creado con éxito!")
                break
    except KeyboardInterrupt:
        MenuPrincipal()
    except:
        print("\n\nHa ocurrido un error.")
        leave = "y"
        return leave
        

def MenuPrincipal():
    print("\n\nBienvenido al menú principal del sistema de Almirante Brown.\n\nElija la opción que desee.")
    try:
        print("   1. Sector de Carga de Trámites.\n   2. Estadisticas del día.\n   3. Salir")
        while True:    
            if keyboard.is_pressed('1'):
                Transit()
                print("\n\n1. Sector de Carga de Trámites.\n2. Estadisticas del día.\n3. Salir")
                continue
            elif keyboard.is_pressed('2'):
                stats()
                print("\n\n1. Sector de Carga de Trámites.\n2. Estadisticas del día.\n3. Salir")
                continue
            elif keyboard.is_pressed('3'):
                print("\nVuelva pronto!")
                leave = "y"
                return leave
            
    except:
        print("Hubo un error! Estamos trabajando en ello.")


def Transit():
    try:
        print("\nBienvenido al sector de Carga de tramites. Para comenzar, introduzca:")
        while True:
            name = input("Nombre y apellido: \n")
            
            namesList.append(name)
            while True:
                dni = input("DNI: \n")
                if (len(dni) > 6) and (len(dni) <= 8) and dni.isdigit():
                    dniList.append(dni)
                    break
                else:
                    print("El dni debe tener entre 6 y 8 digitos")
                    continue
            while True:
                hace_teorico = input("Es ONP? Y/y para si, N/n para no\n").lower()
                if hace_teorico == "y":
                    teorico.append(True)
                    while True:
                        hace_practico = input("Debe hacer el practico? Y/y SI, N/n NO\n").lower()
                        if hace_practico == "y":
                            practico.append(True)
                            break
                        elif hace_practico == "n":
                            practico.append(False)
                            break
                        else:
                            print("No se reconoce la respuesta, recuerde Y/y para SI, N/n para no.\n")
                            continue
                    break
                elif hace_teorico == "n":
                    teorico.append(False)
                    practico.append(False)
                    break
                else:
                    print("No se reconoce la respuesta. Por favor, Y/y para SI, N/n para NO.")
                    continue
            time = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
            data = {
                    "Nombre": namesList,
                    "DNI": dniList,
                    "Teorico": teorico,
                    "Práctico": practico,
                    "Hora del Tramite": time
                }
            newdf = pd.DataFrame(data)
            newone = input("Quieres agregar a alguien mas? Y/y o N/n\n")
            newone = newone.lower()
            while True:
                if newone == "y":
                    print("\n\n", newdf, "\n\n")
                    break
                elif newone == "n":
                    print(newdf, "\n\n")
                    raise KeyboardInterrupt
                    
                else: 
                    print("Error, para si escriba Y o y, para NO, escriba N o n")
            
    except KeyboardInterrupt:
        
        while True:
            save = input("Quiere guardar sus datos? Y/y para SI, N/n para NO\n").lower()
            dia = datetime.today().strftime('%d-%m-%Y')
            if save == "y":
                try:
                    finish = True
                    while True:
                        saved(dia, newdf)
                        if guardado == "yes":
                            finish = False
                            break
                        else:
                            continue
                    if finish == False:
                        break
                    else: 
                        continue
                except PermissionError:
                    print("\nError! Verifique que el archivo \" transito.xlsx\" se encuentre cerrado y que no esté siendo utilizado por otros procesos.\n")
            elif save == "n":
                    sure = input("Está seguro? Todos los datos cargados serán eliminados.\n").lower()
                    if sure == "y":
                        break
                    elif sure == "n":
                        continue
                    else: 
                        print("Input inválido, vuelva a intentarlo")
                        continue
            else:
                print("Por favor, Y/y para guardar, N/n para salir sin guardar")
                continue
    except UnboundLocalError:
        print("Vuelva pronto!")


def stats():
    print("\n\n\n\nProximamente...")


def saved(dia, newdf):

    book = openpyxl.load_workbook(r'C:\Users\Almirante Brown\Saved Games\c\transit-csv\transito.xlsx')

    new_sheet = book.create_sheet(title=dia)
    header = newdf.columns
    for i, h in enumerate(header):
        new_sheet.cell(row=1, column=i+1, value=str(h))
        new_sheet.cell(row=1, column=i+1).font = Font(bold=True)
        new_sheet.cell(row=1, column=i+1).border = Border(bottom=Side(border_style='thin', color='000000'))

    for r in dataframe_to_rows(newdf, index=False):
        new_sheet.append(r)
    book.save(r'C:\Users\Almirante Brown\Saved Games\c\transit-csv\transito.xlsx')
    print("Sus datos han sido guardados! Vuelva pronto!")
    global guardado
    guardado = "yes"

       
UsrEnter()